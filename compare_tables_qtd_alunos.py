# -*- coding: latin-1 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd


df = pd.read_csv('TURMAS.csv', sep='|', usecols=["CO_ENTIDADE", "QT_MATRICULAS"])

newWorkbook = Workbook()
newSheet = newWorkbook.active

# Loads the table file
base_table = load_workbook('tabela_base.xlsx')
inep_qtd_alunos_table = load_workbook('inep_quantidade_alunos.xlsx')

# Now we are working only with Sheet1
base_table_sheet = base_table['Planilha1']
inep_qtd_alunos_table_sheet = inep_qtd_alunos_table['Planilha1']
counter = 1

for iterator_base_table in range(2, 703):

    base_table_codigoINEP = base_table_sheet.cell(row = iterator_base_table, column = 1).value
    newSheet.cell(row=iterator_base_table, column = 1).value = base_table_codigoINEP

    foundEscola = False
    foundEscolaQtdAlunos = 0


    for iterator_inep_qtd_alunos_row in df:
        print(iterator_inep_qtd_alunos_row)
        #print(iterator_inep_qtd_alunos_row.values)
#         codigoINEP_int = iterator_inep_qtd_alunos_row["CO_ENTIDADE"]
#         print(codigoINEP_int)
#
#         if base_table_codigoINEP == codigoINEP_int:
#             foundEscola = True
#             foundEscolaQtdAlunos += qtdAlunos_int
#
#     if foundEscola:
#         newSheet.cell(row=iterator_base_table, column = 2).value = foundEscolaQtdAlunos
#     else:
#         newSheet.cell(row=iterator_base_table, column = 2).value = "FALSE"
#
#     counter += 1
#
#     # print("Counter: " + str(counter))
#     # print("Escola: " + str(base_table_codigoINEP))
#     # print("Quantidade de alunos: " + str(foundEscolaQtdAlunos))
#     # print("----")
#
#
#
#
# newWorkbook.save("tabela_base_qtd_alunos.xlsx")
#     #
#     # numero = str(sentEmails_sheet.cell(row = iteratorSentEmails, column = 1).value)    #get first name from excel for ith row
#     # nome = str(sentEmails_sheet.cell(row = iteratorSentEmails, column = 2).value)    #get first name from excel for ith row
#     # telefone = str(sentEmails_sheet.cell(row = iteratorSentEmails, column = 3).value)    #get first name from excel for ith row
#     # email = str(sentEmails_sheet.cell(row = iteratorSentEmails, column = 4).value)    #get first name from excel for ith row
#     # instituicao = str(sentEmails_sheet.cell(row = iteratorSentEmails, column = 5).value)    #get first name from excel for ith row
#     # codigo = str(sentEmails_sheet.cell(row = iteratorSentEmails, column = 6).value)    #get first name from excel for ith row
#     #
#     # numero2 = str(sentEmails_sheet.cell(row = iteratorSentEmails-1, column = 1).value)    #get first name from excel for ith row
#     # nome2 = str(sentEmails_sheet.cell(row = iteratorSentEmails-1, column = 2).value)    #get first name from excel for ith row
#     # telefone2 = str(sentEmails_sheet.cell(row = iteratorSentEmails-1, column = 3).value)    #get first name from excel for ith row
#     # email2 = str(sentEmails_sheet.cell(row = iteratorSentEmails-1, column = 4).value)    #get first name from excel for ith row
#     # instituicao2 = str(sentEmails_sheet.cell(row = iteratorSentEmails-1, column = 5).value)    #get first name from excel for ith row
#     # codigo2 = str(sentEmails_sheet.cell(row = iteratorSentEmails-1, column = 6).value)    #get first name from excel for ith row
#     #
#     # print("Searched Email: %s" % email)
#     # print("OH YEAH, BITCH. LINHA %d" % (iteratorSentEmails))
#     # print("------")
#     #
#     # if (email != email2):
#     #     ws.cell(row=iteratorSentEmails-1, column = 1).value = numero
#     #     ws.cell(row=iteratorSentEmails-1, column = 2).value = nome
#     #     ws.cell(row=iteratorSentEmails-1, column = 3).value = telefone
#     #     ws.cell(row=iteratorSentEmails-1, column = 4).value = email
#     #     ws.cell(row=iteratorSentEmails-1, column = 5).value = instituicao
#     #     ws.cell(row=iteratorSentEmails-1, column = 6).value = codigo
#     #     wb.save("emailsquejaenviadmos.xlsx")
