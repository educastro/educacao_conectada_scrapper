# -*- coding: latin-1 -*-
from openpyxl import load_workbook
from openpyxl import Workbook

newWorkbook = Workbook()
newSheet = newWorkbook.active

# Loads the table file
base_table = load_workbook('tabela_base.xlsx')
educacao_conectada_reference_table = load_workbook('tabela_referencia_educacao_conectada.xlsx')
pnbl_reference_table = load_workbook('tabela_referencia_pnbl.xlsx')
rural_reference_table = load_workbook('tabela_referencia_rural.xlsx')

# Now we are working only with Sheet1
base_table_sheet = base_table['Planilha1']
educacao_conectada_reference_table_sheet = educacao_conectada_reference_table['Planilha1']
pnbl_reference_table_sheet = pnbl_reference_table['Planilha1']
rural_reference_table_sheet = rural_reference_table['Planilha1']

for iterator_base_table in range(2, 703):

    base_table_codigoINEP = base_table_sheet.cell(row = iterator_base_table, column = 1).value
    base_table_sheet.cell(row=iterator_base_table, column = 1).value = base_table_codigoINEP

    foundEducacaoConectada = False
    foundEducacaoConectadaQtdAlunos = None
    foundEducacaoConectadaVelocidade = 0

    foundPnbl = False
    foundPnblTecnologia = None
    foundPnblVelocidade = 0
    foundPnblData = None

    foundRural = False
    foundRuralTecnologia = None
    foundRuralVelocidade = 0
    foundRuralData = None



    for iterator_educacao_conectada_reference_table_sheet in range(2, 278):
        educacao_conectada_reference_table_codigoINEP = educacao_conectada_reference_table_sheet.cell(row = iterator_educacao_conectada_reference_table_sheet, column = 1).value
        educacao_conectada_reference_table_qtdAlunos = educacao_conectada_reference_table_sheet.cell(row = iterator_educacao_conectada_reference_table_sheet, column = 2).value

        quantidade_de_links = 0

        if base_table_codigoINEP == educacao_conectada_reference_table_codigoINEP:
            foundEducacaoConectada = True
            foundEducacaoConectadaQtdAlunos = educacao_conectada_reference_table_qtdAlunos
            quantidade_de_links += 1
            if educacao_conectada_reference_table_qtdAlunos >= 501:
                foundEducacaoConectadaVelocidade = "100 Mbps"
            elif educacao_conectada_reference_table_qtdAlunos >= 201:
                foundEducacaoConectadaVelocidade = "50 Mbps"
            else:
                foundEducacaoConectadaVelocidade = "20 Mbps"

    for iterator_pnbl_reference_table_sheet in range(2, 580):
        pnbl_reference_table_codigoINEP = pnbl_reference_table_sheet.cell(row = iterator_pnbl_reference_table_sheet, column = 1).value
        pnbl_reference_table_tecnologia = pnbl_reference_table_sheet.cell(row = iterator_pnbl_reference_table_sheet, column = 2).value
        pnbl_reference_table_velocidade = pnbl_reference_table_sheet.cell(row = iterator_pnbl_reference_table_sheet, column = 3).value
        pnbl_reference_table_data_de_ativacao = pnbl_reference_table_sheet.cell(row = iterator_pnbl_reference_table_sheet, column = 4).value

        if base_table_codigoINEP == pnbl_reference_table_codigoINEP:
            foundPnbl = True
            quantidade_de_links += 1
            foundPnblTecnologia = pnbl_reference_table_tecnologia
            foundPnblVelocidade = pnbl_reference_table_velocidade
            foundPnblData = pnbl_reference_table_data_de_ativacao

    for iterator_rural_reference_table_sheet in range(2, 80):
        rural_reference_table_codigoINEP = rural_reference_table_sheet.cell(row = iterator_rural_reference_table_sheet, column = 1).value
        rural_reference_table_tecnologia = rural_reference_table_sheet.cell(row = iterator_rural_reference_table_sheet, column = 2).value
        rural_reference_table_velocidade = rural_reference_table_sheet.cell(row = iterator_rural_reference_table_sheet, column = 3).value
        rural_reference_table_data_de_ativacao = rural_reference_table_sheet.cell(row = iterator_rural_reference_table_sheet, column = 4).value

        if base_table_codigoINEP == rural_reference_table_codigoINEP:
            foundRural = True
            quantidade_de_links += 1
            foundRuralTecnologia = rural_reference_table_tecnologia
            foundRuralVelocidade = rural_reference_table_velocidade
            foundRuralData = rural_reference_table_data_de_ativacao

    base_table_sheet.cell(row=iterator_base_table, column = 3).value = foundEducacaoConectada
    base_table_sheet.cell(row=iterator_base_table, column = 4).value = foundEducacaoConectadaQtdAlunos
    base_table_sheet.cell(row=iterator_base_table, column = 5).value = foundEducacaoConectadaVelocidade
    base_table_sheet.cell(row=iterator_base_table, column = 6).value = foundPnbl
    base_table_sheet.cell(row=iterator_base_table, column = 7).value = foundPnblTecnologia
    base_table_sheet.cell(row=iterator_base_table, column = 8).value = foundPnblVelocidade
    base_table_sheet.cell(row=iterator_base_table, column = 9).value = foundPnblData
    base_table_sheet.cell(row=iterator_base_table, column = 10).value = foundRural
    base_table_sheet.cell(row=iterator_base_table, column = 11).value = foundRuralTecnologia
    base_table_sheet.cell(row=iterator_base_table, column = 12).value = foundRuralVelocidade
    base_table_sheet.cell(row=iterator_base_table, column = 13).value = foundRuralData
    base_table_sheet.cell(row=iterator_base_table, column = 14).value = quantidade_de_links


base_table.save("tabela_base.xlsx")
    #
    # numero = str(sentEmails_sheet.cell(row = iteratorSentEmails, column = 1).value)    #get first name from excel for ith row
    # nome = str(sentEmails_sheet.cell(row = iteratorSentEmails, column = 2).value)    #get first name from excel for ith row
    # telefone = str(sentEmails_sheet.cell(row = iteratorSentEmails, column = 3).value)    #get first name from excel for ith row
    # email = str(sentEmails_sheet.cell(row = iteratorSentEmails, column = 4).value)    #get first name from excel for ith row
    # instituicao = str(sentEmails_sheet.cell(row = iteratorSentEmails, column = 5).value)    #get first name from excel for ith row
    # codigo = str(sentEmails_sheet.cell(row = iteratorSentEmails, column = 6).value)    #get first name from excel for ith row
    #
    # numero2 = str(sentEmails_sheet.cell(row = iteratorSentEmails-1, column = 1).value)    #get first name from excel for ith row
    # nome2 = str(sentEmails_sheet.cell(row = iteratorSentEmails-1, column = 2).value)    #get first name from excel for ith row
    # telefone2 = str(sentEmails_sheet.cell(row = iteratorSentEmails-1, column = 3).value)    #get first name from excel for ith row
    # email2 = str(sentEmails_sheet.cell(row = iteratorSentEmails-1, column = 4).value)    #get first name from excel for ith row
    # instituicao2 = str(sentEmails_sheet.cell(row = iteratorSentEmails-1, column = 5).value)    #get first name from excel for ith row
    # codigo2 = str(sentEmails_sheet.cell(row = iteratorSentEmails-1, column = 6).value)    #get first name from excel for ith row
    #
    # print("Searched Email: %s" % email)
    # print("OH YEAH, BITCH. LINHA %d" % (iteratorSentEmails))
    # print("------")
    #
    # if (email != email2):
    #     ws.cell(row=iteratorSentEmails-1, column = 1).value = numero
    #     ws.cell(row=iteratorSentEmails-1, column = 2).value = nome
    #     ws.cell(row=iteratorSentEmails-1, column = 3).value = telefone
    #     ws.cell(row=iteratorSentEmails-1, column = 4).value = email
    #     ws.cell(row=iteratorSentEmails-1, column = 5).value = instituicao
    #     ws.cell(row=iteratorSentEmails-1, column = 6).value = codigo
    #     wb.save("emailsquejaenviadmos.xlsx")
